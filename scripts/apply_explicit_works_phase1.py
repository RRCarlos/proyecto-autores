from __future__ import annotations

import csv
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill


ROOT = Path(__file__).resolve().parents[1]
CSV_IN = ROOT / "docs" / "TABLA_USOS_AUTORES_LIBROS_I_IV_VALIDACION_POR_LOTES.csv"
XLSX_IN = ROOT / "docs" / "TABLA_USOS_AUTORES_LIBROS_I_IV_VALIDACION_POR_LOTES.xlsx"
CSV_OUT = ROOT / "docs" / "TABLA_USOS_AUTORES_LIBROS_I_IV_OBRAS_EXPLICITAS_FASE1_LIBRO_I.csv"
XLSX_OUT = ROOT / "docs" / "TABLA_USOS_AUTORES_LIBROS_I_IV_OBRAS_EXPLICITAS_FASE1_LIBRO_I.xlsx"
MD_OUT = ROOT / "docs" / "TABLA_USOS_AUTORES_LIBROS_I_IV_OBRAS_EXPLICITAS_FASE1_LIBRO_I.md"


# Fase 1: solo Libro I. Rellenar únicamente cuando el OCR 1617 muestra autor + obra
# en el mismo pasaje y con una forma explícita verificable en facsímil.
EXPLICIT_WORKS = {
    "1": {
        "obra": "Historia natural",
        "evidence": "OCR 1617 líneas 4117-4123: 'Plinio, al fin de su historia natural...'",
    },
    "10": {
        "obra": "Tablas de Beroso",
        "evidence": "OCR 1617 líneas 5810-5815: 'Beroso, cuyas Tabulas poco antes desechamos...'",
    },
    "13": {
        "obra": "Etimologías",
        "evidence": "OCR 1617 líneas 5350-5352: 'Isidoro al fin del libro treze de sus Etymologias...'",
    },
    "20": {
        "obra": "Tablas de Beroso",
        "evidence": "OCR 1617 líneas 5810-5815: 'Beroso, cuyas Tabulas poco antes desechamos...'",
    },
    "35": {
        "obra": "Timeo",
        "evidence": "OCR 1617 línea 7370: 'Platon en el Timeo dize...'",
    },
    "39": {
        "obra": "Itinerario",
        "evidence": "OCR 1617 líneas 7545-7550: 'Antonino en su Itinerario...'",
    },
    "42": {
        "obra": "Historia de los de Fenicia",
        "evidence": "OCR 1617 líneas 7638-7640: 'Philon en la historia de los de Phenicia...'",
    },
}


def esc(value: str) -> str:
    return (value or "").replace("|", "\\|").replace("\n", "<br>").strip()


def main() -> None:
    with CSV_IN.open("r", encoding="utf-8-sig", newline="") as f:
        rows = list(csv.DictReader(f))

    for row in rows:
        row["Obra asociada"] = ""
        if row["Orden"] in EXPLICIT_WORKS:
            item = EXPLICIT_WORKS[row["Orden"]]
            row["Obra asociada"] = item["obra"]
            row["Notas validación 1617"] = (
                row.get("Notas validación 1617", "")
                + f" OBRA EXPLÍCITA — FASE 1 (Libro I): {item['evidence']} Verificar visualmente antes de redacción final."
            )

    headers = list(rows[0].keys())
    with CSV_OUT.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=headers)
        writer.writeheader()
        writer.writerows(rows)

    wb = load_workbook(XLSX_IN)
    ws = wb["Usos autores I-IV"]
    xheaders = [cell.value for cell in ws[1]]
    obra_col = xheaders.index("Obra asociada") + 1
    notas_col = xheaders.index("Notas validación 1617") + 1
    fill = PatternFill("solid", fgColor="70AD47")
    for idx, row in enumerate(rows, start=2):
        ws.cell(row=idx, column=obra_col, value=row["Obra asociada"])
        ws.cell(row=idx, column=notas_col, value=row["Notas validación 1617"])
        if row["Obra asociada"]:
            ws.cell(row=idx, column=obra_col).fill = fill
        for cell in ws[idx]:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    meta = wb["Criterios"] if "Criterios" in wb.sheetnames else wb.create_sheet("Criterios")
    meta.append(["Obras explícitas — fase 1", "Libro I revisado de forma conservadora: solo se rellena obra cuando el OCR 1617 contiene autor + obra en el mismo pasaje. Todas siguen pendientes de verificación visual en facsímil."])
    wb.save(XLSX_OUT)

    lines = [
        "# Obras explícitas — Fase 1: Libro I",
        "",
        "Solo se rellena `Obra asociada` cuando Mariana parece nombrar al autor y la obra en el mismo pasaje del OCR 1617. Todas las entradas siguen pendientes de verificación visual en facsímil.",
        "",
        "| Orden | Libro | Capítulo | Autor citado | Obra asociada | Evidencia / notas |",
        "| --- | --- | --- | --- | --- | --- |",
    ]
    for row in rows:
        if row["Obra asociada"]:
            lines.append(
                "| "
                + " | ".join(
                    [
                        esc(row["Orden"]),
                        esc(row["Libro"]),
                        esc(row["Capítulo"]),
                        esc(row["Autor citado"]),
                        esc(row["Obra asociada"]),
                        esc(EXPLICIT_WORKS[row["Orden"]]["evidence"]),
                    ]
                )
                + " |"
            )
    MD_OUT.write_text("\n".join(lines), encoding="utf-8")

    print(f"created {XLSX_OUT}, {CSV_OUT}, {MD_OUT}")
    print(f"phase 1 explicit works: {len(EXPLICIT_WORKS)}")


if __name__ == "__main__":
    main()
