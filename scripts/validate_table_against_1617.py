from __future__ import annotations

import csv
import re
import unicodedata
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill


ROOT = Path(__file__).resolve().parents[1]
OCR = ROOT / "sources" / "mariana_1617_tomo_primero_ocr.txt"
CSV_IN = ROOT / "docs" / "TABLA_USOS_AUTORES_LIBROS_I_IV.csv"
XLSX_IN = ROOT / "docs" / "TABLA_USOS_AUTORES_LIBROS_I_IV.xlsx"
CSV_OUT = ROOT / "docs" / "TABLA_USOS_AUTORES_LIBROS_I_IV_VALIDACION_POR_LOTES.csv"
XLSX_OUT = ROOT / "docs" / "TABLA_USOS_AUTORES_LIBROS_I_IV_VALIDACION_POR_LOTES.xlsx"


BOOK_RANGES = {
    "I": (3950, 9295),
    "II": (9296, 15335),
    "III": (15336, 21103),
    "IV": (21104, 27084),
}


AUTHOR_VARIANTS = {
    "Plinio el Viejo": ["plinio", "plini"],
    "Solino": ["solino", "folino"],
    "Pomponio Mela": ["pomponio mela", "mela"],
    "Justino": ["justino", "iustino"],
    "Isidoro de Sevilla": ["isidoro", "ifidoro", "ysidoro"],
    "Sexto Rufo / Rufo Festo": ["rufo", "ruso", "festo", "sexto rufo"],
    "Ptolomeo": ["ptolomeo", "tolomeo", "ptolome"],
    "Pseudo-Beroso / Annio de Viterbo": ["beroso", "berofo", "annio", "annio de viterbo"],
    "Apiano de Alejandría": ["apiano", "appiano"],
    "Estrabón": ["estrabon", "eftrabon", "strabon"],
    "Diodoro Sículo": ["diodoro", "diodoro siculo"],
    "Suetonio Tranquilo": ["suetonio", "fuetonio"],
    "Quinto Fabio Píctor": ["fabio pictor", "pictor", "pitor"],
    "Quinto Valerio Sorano": ["valerio sorano", "sorano"],
    "Filisto de Siracusa": ["filisto", "philisto", "filifto"],
    "Aristóteles": ["aristoteles", "ariftoteles"],
    "Platón": ["platon"],
    "Herodoto": ["herodoto"],
    "Anónimo del Itinerarium Antonini": ["antonino", "itinerario"],
    "Lucio Anneo Floro": ["floro", "florian", "floriano"],
    "Filón de Biblos": ["filon", "philon", "biblos"],
    "Silio Itálico": ["silio", "filio italico", "silio italico"],
    "Flavio Josefo": ["josefo", "iosepho", "iòsepho", "josepho"],
    "Tito Livio": ["livio", "liuio"],
    "Polibio": ["polibio", "polybio"],
    "Festo Pompeyo": ["festo pompeyo", "pompeyo festo", "festo"],
    "Arriano": ["arriano", "arriano de nicomedia"],
    "Tucídides": ["tucidides", "thucidides"],
    "Dionisio de Halicarnaso": ["dionisio", "halicarnaso"],
    "Hipócrates de Cos": ["hipocrates", "hippocrates"],
    "Paulo Orosio": ["orosio", "orofio"],
    "Eusebio de Cesarea": ["eusebio", "eufebio"],
    "Plutarco": ["plutarco", "plutarcho"],
    "Valerio Máximo": ["valerio maximo"],
    "Aulo Gelio": ["aulo gelio", "gelio", "gellio"],
    "Cicerón": ["ciceron", "ciceron"],
    "Andrea de' Bardi": ["anconitano", "andrea", "bardi"],
    "Antonio de Nebrija": ["nebrija", "antonio de nebrija"],
    "Casiodoro": ["casiodoro", "cassiodoro"],
    "Tertuliano": ["tertuliano", "tertulliano"],
    "Ambrosio de Morales": ["ambrosio de morales", "morales"],
    "Lucano": ["lucano", "lucan"],
    "Virgilio": ["virgilio", "vergilio"],
    "Dion Casio": ["dion", "dio", "dion casio"],
    "Joan Margarit i Pau": ["margarite", "margarit"],
    "Hilderico": ["hilderico", "ilderico"],
    "Ahmad al-Razi / el moro Rasis": ["rasis", "razi", "moro rasis"],
    "Agustín de Hipona": ["agustin", "augustin"],
    "Tácito": ["tacito", "tacitus"],
    "Lucas evangelista": ["lucas", "san lucas"],
    "Séneca el Viejo": ["seneca", "feneca"],
    "Quintiliano": ["quintiliano", "quintilian"],
    "Pelayo de Oviedo": ["pelayo"],
    "Beda el Venerable": ["beda"],
    "Usuardo": ["usuardo", "ufuardo"],
    "Poggio Bracciolini": ["poggio", "pogio"],
    "Pietro Crinito": ["crinito", "prieto crinito", "pedro crinito"],
    "Lilio Gregorio Giraldi": ["giraldo", "giraldi", "gregorio giraldo"],
    "Miguel Sincelo": ["sincelo", "syncelo", "miguel sincelo"],
    "Julio Capitolino": ["capitolino", "julio capitolino"],
    "Nicéforo": ["niceforo", "nicephoro"],
    "Trebellio Polión": ["trebellio", "polion"],
    "Códices de Teodosio y Justiniano": ["teodosio", "justiniano", "codigo"],
    "Prudencio": ["prudencio"],
    "Braulio de Zaragoza": ["braulio"],
    "Jerónimo": ["jeronimo", "hieronymo"],
    "Amiano Marcelino": ["amiano", "marcelino", "marcellino"],
    "Sulpicio Severo": ["sulpicio", "severo sulpicio"],
    "Sigeberto de Gembloux": ["sigiberto", "sigeberto"],
    "Claudiano": ["claudiano"],
    "Ildefonso de Toledo": ["ildefonso", "ildephonso"],
    "Hernando del Pulgar": ["pulgar", "hernando del pulgar"],
    "Próspero de Aquitania": ["prospero", "aquitanio"],
    "Juan de Mariana": ["mariana", "historiae de rebus"],
    "Escritura del monasterio de San Clemente de Toledo": ["san clemente", "monasterio de san clemente"],
}


def norm(text: str) -> str:
    text = unicodedata.normalize("NFD", text)
    text = "".join(c for c in text if unicodedata.category(c) != "Mn")
    text = text.lower()
    text = text.replace("ſ", "s")
    text = re.sub(r"[^a-z0-9ñ]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def soft_norm(text: str) -> str:
    value = norm(text)
    # Common OCR/early-modern type confusion: long s appears as f.
    return value.replace("f", "s").replace("v", "u")


def line_number_from_offset(line_offsets: list[int], offset: int) -> int:
    lo, hi = 0, len(line_offsets) - 1
    while lo <= hi:
        mid = (lo + hi) // 2
        if line_offsets[mid] <= offset:
            lo = mid + 1
        else:
            hi = mid - 1
    return max(1, hi + 1)


def find_variant(segment_norm: str, variants: list[str]) -> tuple[bool, str]:
    for variant in variants:
        v = norm(variant)
        if v and v in segment_norm:
            return True, variant
        sv = soft_norm(variant)
        if sv and sv in soft_norm(segment_norm):
            return True, variant
    return False, ""


def find_ocr_candidate_snippet(lines: list[str], start: int, end: int, variants: list[str]) -> tuple[str, str]:
    """Return an OCR candidate snippet around the matched author name.

    This is NOT a verified quotation. It is only a navigation aid for visual
    checking against the facsimile.
    """
    for idx in range(max(1, start), min(len(lines), end) + 1):
        line = lines[idx - 1]
        line_norm = norm(line)
        line_soft = soft_norm(line_norm)
        for variant in variants:
            if norm(variant) in line_norm or soft_norm(variant) in line_soft:
                before = lines[max(0, idx - 3) : idx - 1]
                after = lines[idx : min(len(lines), idx + 2)]
                snippet_lines = before + [line] + after
                snippet = " / ".join(s.strip() for s in snippet_lines if s.strip())
                snippet = re.sub(r"\s+", " ", snippet).strip()
                return snippet, str(idx)
    return "", ""


def build_chapter_ranges(lines: list[str]) -> dict[tuple[str, str], tuple[int, int]]:
    # Known actual OCR starts for reliable segmentation. Missing chapters inherit a broad
    # range from previous/next known chapter within the same book.
    starts = {
        ("I", "I"): 3958,
        ("I", "III"): 4358,
        ("I", "V"): 4914,
        ("I", "VI"): 5030,
        ("I", "VII"): 5112,
        ("I", "XI"): 6133,
        ("I", "XIII"): 6728,
        ("I", "XIV"): 6990,
        ("I", "XV"): 7203,
        ("I", "XVI"): 7453,
        ("I", "XVII"): 7665,
        ("I", "XX"): 8602,
        ("I", "XXII"): 9070,
        ("II", "I"): 9296,
        ("II", "II"): 9459,
        ("II", "V"): 10127,
        ("II", "VI"): 10451,
        ("II", "VII"): 10742,
        ("II", "VIII"): 10941,
        ("II", "X"): 11651,
        ("II", "XVII"): 12954,
        ("II", "XVIII"): 13132,
        ("II", "XIX"): 13342,
        ("II", "XXIII"): 14182,
        ("II", "XXV"): 14710,
        ("III", "I"): 15336,
        ("III", "III"): 16077,
        ("III", "VI"): 16760,
        ("III", "X"): 17449,
        ("III", "XII"): 18125,
        ("III", "XVI"): 18934,
        ("III", "XVIII"): 19325,
        ("III", "XXIV"): 20074,
        ("III", "XXV"): 20550,
        ("IV", "I"): 21104,
        ("IV", "III"): 21757,
        ("IV", "IV"): 22217,
        ("IV", "VIII"): 23418,
        ("IV", "X"): 23883,
        ("IV", "XII"): 24368,
        ("IV", "XIII"): 24680,
        ("IV", "XIV"): 24876,
        ("IV", "XXI"): 26395,
    }

    all_chapters = {
        "I": ["I", "II", "III", "IV", "V", "VI", "VII", "VIII", "IX", "X", "XI", "XII", "XIII", "XIV", "XV", "XVI", "XVII", "XVIII", "XIX", "XX", "XXI", "XXII"],
        "II": ["I", "II", "III", "IV", "V", "VI", "VII", "VIII", "IX", "X", "XI", "XII", "XIII", "XIV", "XV", "XVI", "XVII", "XVIII", "XIX", "XX", "XXI", "XXII", "XXIII", "XXIV", "XXV", "XXVI"],
        "III": ["I", "II", "III", "IV", "V", "VI", "VII", "VIII", "IX", "X", "XI", "XII", "XIII", "XIV", "XV", "XVI", "XVII", "XVIII", "XX", "XXI", "XXII", "XXIII", "XXIV", "XXV"],
        "IV": ["I", "II", "III", "IV", "V", "VI", "VII", "VIII", "IX", "X", "XI", "XII", "XIII", "XIV", "XV", "XVI", "XVII", "XVIII", "XIX", "XX", "XXI"],
    }
    ranges: dict[tuple[str, str], tuple[int, int]] = {}
    for book, chapters in all_chapters.items():
        book_start, book_end = BOOK_RANGES[book]
        for idx, chapter in enumerate(chapters):
            key = (book, chapter)
            if key in starts:
                start = starts[key]
            else:
                prev = [starts[(book, c)] for c in chapters[:idx] if (book, c) in starts]
                start = prev[-1] if prev else book_start
            next_starts = [starts[(book, c)] for c in chapters[idx + 1 :] if (book, c) in starts]
            end = (next_starts[0] - 1) if next_starts else book_end
            ranges[key] = (start, end)
    return ranges


def review_batch(book: str, chapter: str) -> str:
    roman_to_int = {
        "I": 1,
        "II": 2,
        "III": 3,
        "IV": 4,
        "V": 5,
        "VI": 6,
        "VII": 7,
        "VIII": 8,
        "IX": 9,
        "X": 10,
        "XI": 11,
        "XII": 12,
        "XIII": 13,
        "XIV": 14,
        "XV": 15,
        "XVI": 16,
        "XVII": 17,
        "XVIII": 18,
        "XIX": 19,
        "XX": 20,
        "XXI": 21,
        "XXII": 22,
        "XXIII": 23,
        "XXIV": 24,
        "XXV": 25,
        "XXVI": 26,
    }
    n = roman_to_int.get(chapter, 0)
    if book == "I":
        if n <= 7:
            return "Lote 1 — Libro I, caps. I-VII"
        if n <= 15:
            return "Lote 2 — Libro I, caps. VIII-XV"
        return "Lote 3 — Libro I, caps. XVI-XXII"
    if book == "II":
        if n <= 10:
            return "Lote 4 — Libro II, caps. I-X"
        if n <= 20:
            return "Lote 5 — Libro II, caps. XI-XX"
        return "Lote 6 — Libro II, caps. XXI-XXVI"
    if book == "III":
        if n <= 10:
            return "Lote 7 — Libro III, caps. I-X"
        if n <= 18:
            return "Lote 8 — Libro III, caps. XI-XVIII"
        return "Lote 9 — Libro III, caps. XX-XXV"
    if book == "IV":
        if n <= 7:
            return "Lote 10 — Libro IV, caps. I-VII"
        if n <= 14:
            return "Lote 11 — Libro IV, caps. VIII-XIV"
        return "Lote 12 — Libro IV, caps. XV-XXI"
    return "Sin lote"


def validate() -> None:
    text = OCR.read_text(encoding="utf-8", errors="ignore")
    lines = text.splitlines()
    line_offsets = []
    pos = 0
    for line in lines:
        line_offsets.append(pos)
        pos += len(line) + 1

    chapter_ranges = build_chapter_ranges(lines)

    with CSV_IN.open("r", encoding="utf-8-sig", newline="") as f:
        rows = list(csv.DictReader(f))
    base_headers = list(rows[0].keys())
    extra = [
        "Lote de revisión",
        "Validación 1617",
        "Coincidencia OCR 1617",
        "Fragmento OCR candidato 1617",
        "Línea OCR candidata 1617",
        "Rango OCR 1617",
        "Transcripción verificada manualmente",
        "Estado transcripción",
        "Notas validación 1617",
    ]

    output = []
    for row in rows:
        book = row["Libro"]
        chapter = row["Capítulo"]
        norm_name = row["Forma normalizada"]
        variants = AUTHOR_VARIANTS.get(norm_name, [norm_name, row["Autor citado"]])
        start, end = chapter_ranges.get((book, chapter), BOOK_RANGES[book])
        # Expand uncertain inherited ranges a little, but keep inside the book.
        bstart, bend = BOOK_RANGES[book]
        start = max(bstart, start - 15)
        end = min(bend, end + 15)
        segment = "\n".join(lines[start - 1 : end])
        segment_norm = norm(segment)
        ok, variant = find_variant(segment_norm, variants)
        citation, citation_line = "", ""
        if ok:
            status = "Confirmado en 1617"
            match = variant
            citation, citation_line = find_ocr_candidate_snippet(lines, start, end, variants)
            note = "Nombre localizado en el rango OCR del mismo libro/capítulo aproximado. El fragmento OCR NO es cita validada; revisar facsímil."
        else:
            book_segment = "\n".join(lines[bstart - 1 : bend])
            ok_book, variant_book = find_variant(norm(book_segment), variants)
            if ok_book:
                status = "Localizado en el libro; capítulo no confirmado por OCR"
                match = variant_book
                citation, citation_line = find_ocr_candidate_snippet(lines, bstart, bend, variants)
                note = "El OCR localiza el nombre dentro del mismo libro, pero no dentro del rango capitular aproximado. Requiere revisión manual."
            else:
                status = "No localizado automáticamente en 1617"
                match = ""
                citation, citation_line = "", ""
                note = "No hay coincidencia automática fiable; puede deberse a OCR, variante nominal o error de las notas. Requiere revisión manual."

        row.update(
            {
                "Lote de revisión": review_batch(book, chapter),
                "Validación 1617": status,
                "Coincidencia OCR 1617": match,
                "Fragmento OCR candidato 1617": citation,
                "Línea OCR candidata 1617": citation_line,
                "Rango OCR 1617": f"líneas {start}-{end}",
                "Transcripción verificada manualmente": "",
                "Estado transcripción": "Pendiente de verificación visual",
                "Notas validación 1617": note,
            }
        )
        output.append(row)

    headers = base_headers + extra
    with CSV_OUT.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=headers)
        writer.writeheader()
        writer.writerows(output)

    # Update spreadsheet preserving wrapped formatting.
    wb = load_workbook(XLSX_IN)
    ws = wb["Usos autores I-IV"]
    start_col = ws.max_column + 1
    fill = PatternFill("solid", fgColor="7030A0")
    for idx, header in enumerate(extra, start=start_col):
        cell = ws.cell(row=1, column=idx, value=header)
        cell.fill = fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
    for r_idx, row in enumerate(output, start=2):
        for c_offset, header in enumerate(extra):
            cell = ws.cell(row=r_idx, column=start_col + c_offset, value=row[header])
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    widths = [32, 28, 28, 78, 18, 24, 78, 30, 62]
    for i, width in enumerate(widths, start=start_col):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = width
    ws.auto_filter.ref = ws.dimensions

    meta = wb["Criterios"] if "Criterios" in wb.sheetnames else wb.create_sheet("Criterios")
    meta.append(["Validación 1617", "Automática por OCR de Internet Archive/Google Books identifier bub_gb_qpHbv84HpA0C. El fragmento OCR candidato NO es cita literal fiable: solo sirve para ubicar el pasaje. La cita académica debe escribirse en 'Transcripción verificada manualmente' tras comprobar el facsímil."])
    wb.save(XLSX_OUT)

    counts = {}
    for row in output:
        counts[row["Validación 1617"]] = counts.get(row["Validación 1617"], 0) + 1
    print(f"created {XLSX_OUT} and {CSV_OUT}")
    print(counts)


if __name__ == "__main__":
    validate()
