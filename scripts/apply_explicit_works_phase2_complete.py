from __future__ import annotations

import csv
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill


ROOT = Path(__file__).resolve().parents[1]
CSV_IN = ROOT / "docs" / "TABLA_USOS_AUTORES_LIBROS_I_IV_VALIDACION_POR_LOTES.csv"
XLSX_IN = ROOT / "docs" / "TABLA_USOS_AUTORES_LIBROS_I_IV_VALIDACION_POR_LOTES.xlsx"
CSV_OUT = ROOT / "docs" / "TABLA_USOS_AUTORES_LIBROS_I_IV_OBRAS_EXPLICITAS_COMPLETO.csv"
XLSX_OUT = ROOT / "docs" / "TABLA_USOS_AUTORES_LIBROS_I_IV_OBRAS_EXPLICITAS_COMPLETO.xlsx"
MD_OUT = ROOT / "docs" / "TABLA_USOS_AUTORES_LIBROS_I_IV_OBRAS_EXPLICITAS_COMPLETO.md"

# Obras explícitas — TODOS los libros (I-IV)
# Solo se rellena "Obra asociada" cuando el OCR 1617 contiene autor + obra
# en el mismo pasaje con forma explícita verificable en facsímil.
# Cada entrada tiene: "obra" y "evidence" (localización en OCR 1617).
EXPLICIT_WORKS = {
    # ── LIBRO I (Phase 1, ya confirmado) ──
    "1": {
        "obra": "Historia natural",
        "evidence": "OCR 1617 líneas 4117-4123: \"Plinio, al fin de su historia natural...\"",
    },
    "10": {
        "obra": "Tablas de Beroso",
        "evidence": "OCR 1617 líneas 5810-5815: \"Beroso, cuyas Tabulas poco antes desechamos...\"",
    },
    "13": {
        "obra": "Etimologías",
        "evidence": "OCR 1617 líneas 5350-5352: \"Isidoro al fin del libro treze de sus Etymologias...\"",
    },
    "20": {
        "obra": "Tablas de Beroso",
        "evidence": "OCR 1617 líneas 5810-5815: \"Beroso, cuyas Tabulas poco antes desechamos...\"",
    },
    "35": {
        "obra": "Timeo",
        "evidence": "OCR 1617 línea 7370: \"Platon en el Timeo dize...\"",
    },
    "39": {
        "obra": "Itinerario",
        "evidence": "OCR 1617 líneas 7545-7550: \"Antonino en su Itinerario...\"",
    },
    "42": {
        "obra": "Historia de los de Fenicia",
        "evidence": "OCR 1617 líneas 7638-7640: \"Philon en la historia de los de Phenicia...\"",
    },

    # ── LIBRO III ──
    "109": {
        "obra": "Égloga",
        "evidence": "OCR 1617 línea 19995-19996: \"do por vna Egloga de Virgilio, en que...\"",
    },

    # ── LIBRO IV ──
    "132": {
        "obra": "Instituciones oratorias",
        "evidence": "OCR 1617 líneas 22077-22081: \"trofi a Fauio Quintiliano... Sus inftituciones oratorias eftuuieron perdidas...\"",
    },
    "140": {
        "obra": "Instituciones oratorias",
        "evidence": "OCR 1617 líneas 22077-22081: misma referencia que orden 132: \"Sus inftituciones oratorias eftuuieron perdidas...\"",
    },
    "144": {
        "obra": "Instituciones oratorias",
        "evidence": "OCR 1617 líneas 22077-22081: misma referencia que orden 132: \"Sus inftituciones oratorias eftuuieron perdidas...\"",
    },
    "134": {
        "obra": "De viris illustribus / Varones ilustres",
        "evidence": "OCR 1617 líneas 25844-25847: \"fan Ifidoro en los varones iluftres tomó lo que queda dicho.\"",
    },
    "168": {
        "obra": "De viris illustribus / Varones ilustres",
        "evidence": "OCR 1617 líneas 25844-25847: \"fan Ifidoro en los varones iluftres tomó lo que queda dicho.\"",
    },
    "172": {
        "obra": "De viris illustribus / Varones ilustres",
        "evidence": "OCR 1617 líneas 25844-25847: \"fan Ifidoro en los varones iluftres tomó lo que queda dicho.\"",
    },
    "174": {
        "obra": "De viris illustribus / Varones ilustres",
        "evidence": "OCR 1617 líneas 25844-25847: \"fan Ifidoro en los varones iluftres tomó lo que queda dicho.\"",
    },
    "175": {
        "obra": "De viris illustribus / Varones ilustres",
        "evidence": "OCR 1617 líneas 25844-25847: \"fan Ifidoro en los varones iluftres tomó lo que queda dicho.\"",
    },
    "153": {
        "obra": "Código de Justiniano / Código de Teodosio",
        "evidence": "OCR 1617 línea 22870-22871: \"algunas leyes del Codigo de Iustinianp\"; línea 25410-25411: \"aun en el Codigo de Teodofio...\"",
    },
    "155": {
        "obra": "Vida de Santa Leocadia",
        "evidence": "OCR 1617 líneas 24783-24787: \"la historia de la vida y muerte de Santa Leocadia... escrita por Braulio Obispo de Zaragoça...\"",
    },
}


def esc(value: str) -> str:
    return (value or "").replace("|", "\\|").replace("\n", "<br>").strip()


def main() -> None:
    with CSV_IN.open("r", encoding="utf-8-sig", newline="") as f:
        rows = list(csv.DictReader(f))

    count_new = 0
    for row in rows:
        # Preserve existing obra asociada from Phase 1
        row["Obra asociada"] = row.get("Obra asociada", "") or ""
        if row["Orden"] in EXPLICIT_WORKS and not row["Obra asociada"]:
            item = EXPLICIT_WORKS[row["Orden"]]
            row["Obra asociada"] = item["obra"]
            prev_notes = row.get("Notas validación 1617", "") or ""
            sep = " " if prev_notes else ""
            row["Notas validación 1617"] = (
                prev_notes
                + f"{sep}OBRA EXPLÍCITA — FASE 2 (completo): {item['evidence']} Verificar visualmente antes de redacción final."
            )
            count_new += 1

    headers = list(rows[0].keys())
    with CSV_OUT.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=headers)
        writer.writeheader()
        writer.writerows(rows)

    wb = load_workbook(XLSX_IN)
    ws = wb["Usos autores I-IV"]
    xheaders = [cell.value for cell in ws[1]]
    obra_col = xheaders.index("Obra asociada") + 1
    notas_col = xheaders.index("Notas validación 1617") + 1
    fill = PatternFill("solid", fgColor="70AD47")
    for idx, row in enumerate(rows, start=2):
        ws.cell(row=idx, column=obra_col, value=row["Obra asociada"])
        ws.cell(row=idx, column=notas_col, value=row["Notas validación 1617"])
        if row["Obra asociada"]:
            ws.cell(row=idx, column=obra_col).fill = fill
        for cell in ws[idx]:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    meta = wb["Criterios"] if "Criterios" in wb.sheetnames else wb.create_sheet("Criterios")
    meta.append(["Obras explícitas — completo (Fase 1 + 2)",
                 "Libros I-IV revisados de forma conservadora: solo se rellena obra cuando el OCR 1617 contiene autor + obra en el mismo pasaje. "
                 "Todas siguen pendientes de verificación visual en facsímil."])
    wb.save(XLSX_OUT)

    # ── Markdown ──
    lines = [
        "# Obras explícitas — Completo: Libros I-IV",
        "",
        "Solo se rellena `Obra asociada` cuando Mariana nombra al autor **y** la obra en el mismo pasaje del OCR 1617. "
        "Todas las entradas siguen pendientes de verificación visual en facsímil.",
        "",
        "| Orden | Libro | Capítulo | Autor citado | Obra asociada | Evidencia (OCR 1617) |",
        "| --- | --- | --- | --- | --- | --- |",
    ]
    for row in rows:
        if row["Obra asociada"] and row["Orden"] in EXPLICIT_WORKS:
            item = EXPLICIT_WORKS[row["Orden"]]
            evidence = item["evidence"]
        elif row["Obra asociada"]:
            evidence = "Verificar visualmente — evidencia pendiente de registro."
        else:
            continue
        lines.append(
            "| "
            + " | ".join([
                esc(row["Orden"]),
                esc(row["Libro"]),
                esc(row["Capítulo"]),
                esc(row["Autor citado"]),
                esc(row["Obra asociada"]),
                esc(evidence),
            ])
            + " |"
        )

    MD_OUT.write_text("\n".join(lines), encoding="utf-8")

    # Summary
    assigned_rows = [r for r in rows if r["Obra asociada"]]
    by_book: dict[str, int] = {}
    for r in assigned_rows:
        by_book[r["Libro"]] = by_book.get(r["Libro"], 0) + 1

    print(f"Archivos creados:")
    print(f"  {XLSX_OUT.name}")
    print(f"  {CSV_OUT.name}")
    print(f"  {MD_OUT.name}")
    print()
    print(f"Total filas con obra explícita: {len(assigned_rows)}")
    print(f"Nuevas en Fase 2: {count_new}")
    for bk, cnt in sorted(by_book.items()):
        print(f"  {bk}: {cnt} filas")
    print()
    for r in assigned_rows:
        print(f"  Orden {r['Orden']:>3s} | {r['Libro']} | {r['Autor citado']:<30s} → {r['Obra asociada']}")


if __name__ == "__main__":
    main()
