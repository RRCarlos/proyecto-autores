from __future__ import annotations

import csv
import html
import re
import zipfile
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


ROOT = Path(__file__).resolve().parents[1]
DOCX = Path(r"C:\Users\PC\Downloads\Historia general de España.docx")
OUT_XLSX = ROOT / "docs" / "TABLA_USOS_AUTORES_LIBROS_I_IV.xlsx"
OUT_CSV = ROOT / "docs" / "TABLA_USOS_AUTORES_LIBROS_I_IV.csv"


BOOKS = {
    "Primero": "I",
    "Segundo": "II",
    "Tercero": "III",
    "Cuarto": "IV",
}


AUTHORS = [
    # alias regex, normalized name, associated work, source type
    (r"Cayo\s+Plinio\s+Segundo|Plinio", "Plinio el Viejo", "Naturalis Historia", "Clásica romana"),
    (r"Cayo\s+Julio\s+Solino|Solino", "Solino", "Collectanea rerum memorabilium", "Clásica romana / compilatoria"),
    (r"Pomponio\s+Mela|\bMela\b", "Pomponio Mela", "De chorographia / De situ orbis", "Clásica romana"),
    (r"Marco\s+Juniano\s+Justino|Justino", "Justino", "Epitoma historiarum Philippicarum", "Clásica romana / epítome"),
    (r"San\s+Isidoro|San\s+Idisoro|Isidoro\s+de\s+Sevilla", "Isidoro de Sevilla", "Etymologiae / obras históricas", "Medieval / patrística hispana"),
    (r"Sexto\s+Rufo|Rufo\s+Festo", "Sexto Rufo / Rufo Festo", "Breviarium", "Tardorromana / compendio"),
    (r"Claudio\s+Ptolomeo|Ptolomeo", "Ptolomeo", "Geographia", "Clásica grecorromana"),
    (r"Falso\s+Beroso|pseudo\s+Beroso|Beroso", "Pseudo-Beroso / Annio de Viterbo", "Antiquitatum variarum volumina XVII", "Falsificación humanista"),
    (r"Apiano|Appiano", "Apiano de Alejandría", "Historia romana, Iberiké", "Clásica griega / transmisión fragmentaria"),
    (r"Estrabón", "Estrabón", "Geographica", "Clásica griega"),
    (r"Diodoro\s+(?:de\s+Sicilia|Siculo|Sículo)", "Diodoro Sículo", "Bibliotheca historica", "Clásica griega / compilatoria"),
    (r"Suetonio", "Suetonio Tranquilo", "De vita Caesarum", "Clásica romana"),
    (r"Quinto\s+Fabio\s+Píctor|Fabio\s+Píctor", "Quinto Fabio Píctor", "Annales / obra perdida", "Clásica romana / obra perdida"),
    (r"Quinto\s+Valerio\s+Sorano|Valerio\s+Sorano", "Quinto Valerio Sorano", "No determinada", "Clásica romana"),
    (r"Filisto", "Filisto de Siracusa", "Historias sicilianas", "Griega / obra perdida"),
    (r"Aristóteles", "Aristóteles", "No determinada", "Clásica griega"),
    (r"Platón", "Platón", "Timeo", "Clásica griega"),
    (r"Herodoto", "Herodoto", "Historias", "Clásica griega"),
    (r"Antonino", "Anónimo del Itinerarium Antonini", "Itinerarium Antonini Augusti", "Tardorromana / itinerario"),
    (r"Florian|Floro", "Lucio Anneo Floro", "Epitome de Tito Livio", "Clásica romana / epítome"),
    (r"Filón\s+de\s+Biblos", "Filón de Biblos", "Historia fenicia", "Clásica / fragmentaria"),
    (r"Cayo\s+Silio\s+Itálico|Silio\s+Itálico", "Silio Itálico", "Punica", "Clásica romana / poesía épica"),
    (r"Flavio\s+Josefo|Tito\s+Flavio\s+Josefo", "Flavio Josefo", "Antigüedades judías", "Judeo-romana"),
    (r"Tito\s+Livio|Tito\s+Libio", "Tito Livio", "Ab urbe condita", "Clásica romana"),
    (r"Polybio|Polibio", "Polibio", "Historias", "Clásica griega"),
    (r"Festo\s+Pompeyo", "Festo Pompeyo", "De verborum significatu", "Clásica romana / anticuaria"),
    (r"Lucio\s+Flavio\s+Arriano|Arriano", "Arriano", "Obra histórica/geográfica", "Clásica griega"),
    (r"Tucídides", "Tucídides", "Historia de la guerra del Peloponeso", "Clásica griega"),
    (r"Dionisio\s+de\s+Halicarnaso", "Dionisio de Halicarnaso", "Antigüedades romanas", "Clásica griega"),
    (r"Hipócrates", "Hipócrates de Cos", "Corpus Hippocraticum", "Clásica griega / médica"),
    (r"Paulo\s+Orosio|Orosio", "Paulo Orosio", "Historiae adversus paganos", "Tardorromana cristiana"),
    (r"Eusebio(?:\s+Cesariense|\s+de\s+Cesarea)?", "Eusebio de Cesarea", "Chronicon / Historia ecclesiastica", "Cristiana tardoantigua"),
    (r"Plutarco", "Plutarco", "Vidas paralelas", "Clásica griega"),
    (r"Valerio\s+Máximo", "Valerio Máximo", "Factorum et dictorum memorabilium", "Clásica romana"),
    (r"Aulus\s+Gellius|Aulo\s+Gelio", "Aulo Gelio", "Noctes Atticae", "Clásica romana"),
    (r"Marco\s+Tulio\s+Cicerón|Cicerón", "Cicerón", "No determinada", "Clásica romana"),
    (r"Andrea\s+de'?\s*Bardi|Anconitano", "Andrea de' Bardi", "No determinada", "Dudosa"),
    (r"Antonio\s+de\s+Nebrija", "Antonio de Nebrija", "Decades rerum Hispanicarum", "Humanista española"),
    (r"Casiodoro", "Casiodoro", "Chronica", "Tardorromana / medieval temprana"),
    (r"Tertuliano|Tertulliano", "Tertuliano", "No determinada", "Patrística latina"),
    (r"Ambrosio\s+de\s+Morales", "Ambrosio de Morales", "Crónica general de España", "Humanista española"),
    (r"Marco\s+Anneo\s+Lucano|Lucano", "Lucano", "Pharsalia / Bellum civile", "Clásica romana / poesía épica"),
    (r"Virgilio", "Virgilio", "Églogas", "Clásica romana / poesía"),
    (r"Dion\s+Casio|\bDion\b", "Dion Casio", "Historia romana", "Clásica griega/romana"),
    (r"Juan\s+Margarite|Joan\s+Margarit", "Joan Margarit i Pau", "Paralipomenon Hispaniae", "Humanista / eclesiástica hispana"),
    (r"Hilderico", "Hilderico", "No determinada", "Pendiente de identificación"),
    (r"moro\s+Rasis|Rasis|al-Razi|al-Rāzī", "Ahmad al-Razi / el moro Rasis", "Crónica / tradición andalusí", "Andalusí / medieval"),
    (r"San\s+Agustín", "Agustín de Hipona", "No determinada", "Patrística latina"),
    (r"Tácito", "Tácito", "Annales", "Clásica romana"),
    (r"San\s+Lucas", "Lucas evangelista", "Hechos de los Apóstoles", "Bíblica"),
    (r"Seneca\s+el\s+Viejo|Séneca\s+el\s+Viejo", "Séneca el Viejo", "No determinada", "Clásica romana"),
    (r"Quintiliano", "Quintiliano", "Institutio oratoria", "Clásica romana"),
    (r"Pelayo", "Pelayo de Oviedo", "No determinada", "Medieval hispana"),
    (r"Beda", "Beda el Venerable", "No determinada", "Medieval cristiana"),
    (r"Usuardo", "Usuardo", "Martyrologium", "Medieval cristiana"),
    (r"Poggio", "Poggio Bracciolini", "No determinada", "Humanista italiana"),
    (r"Prieto\s+Crinito|Pietro\s+Ricci|Crinito", "Pietro Crinito", "No determinada", "Humanista italiana"),
    (r"Gregorio\s+Giraldo|Giraldo", "Lilio Gregorio Giraldi", "No determinada", "Humanista italiana"),
    (r"Miguel\s+Sincelo", "Miguel Sincelo", "Vida de San Clemente", "Bizantina / manuscrita"),
    (r"Julio\s+Capitolino", "Julio Capitolino", "Historia Augusta", "Historia Augusta / autoría dudosa"),
    (r"Nicéforo", "Nicéforo", "No determinada", "Bizantina cristiana"),
    (r"Trebellio\s+Polión", "Trebellio Polión", "Historia Augusta", "Historia Augusta / autoría dudosa"),
    (r"Códices\s+de\s+Teodosio\s+y\s+Justiniano|Teodosio\s+y\s+Justiniano", "Códices de Teodosio y Justiniano", "Codex Theodosianus / Codex Justinianus", "Jurídica tardoantigua"),
    (r"Prudencio", "Prudencio", "Peristephanon", "Cristiana hispanorromana"),
    (r"Braulio", "Braulio de Zaragoza", "Vida de Santa Leocadia", "Medieval / eclesiástica hispana"),
    (r"San\s+Jerónimo", "Jerónimo", "No determinada", "Patrística latina"),
    (r"Amiano\s+Marcellino|Marcellino", "Amiano Marcelino", "Res gestae", "Tardorromana"),
    (r"Severo\s+Sulpicio|Sulpicio\s+Severo", "Sulpicio Severo", "Chronica", "Cristiana tardoantigua"),
    (r"Sigiberto|Sigeberto", "Sigeberto de Gembloux", "Chronographia", "Medieval cristiana"),
    (r"Claudiano", "Claudiano", "No determinada", "Tardorromana / poesía"),
    (r"San\s+Ildefonso", "Ildefonso de Toledo", "No determinada", "Medieval / eclesiástica hispana"),
    (r"Hernando\s+del\s+Pulgar", "Hernando del Pulgar", "Claros varones de Castilla", "Humanista castellana"),
    (r"Próspero\s+Aquitanio", "Próspero de Aquitania", "Chronicon", "Cristiana tardoantigua"),
    (r"Juan\s+de\s+Mariana|historia\s+Latina|Historiae\s+de\s+rebus\s+Hispaniae", "Juan de Mariana", "Historiae de rebus Hispaniae", "Autor moderno / texto propio"),
    (r"Escritura\s+en\s+el\s+monasterio\s+de\s+San\s+Clemente", "Escritura del monasterio de San Clemente de Toledo", "Documento no identificado", "Documental / archivo"),
]


def extract_paragraphs(docx: Path) -> list[str]:
    xml = zipfile.ZipFile(docx).read("word/document.xml").decode("utf-8")
    raw_paragraphs = re.split(r"</w:p>", xml)
    paragraphs: list[str] = []
    for raw in raw_paragraphs:
        raw = re.sub(r"<w:tab[^>]*/>", "\t", raw)
        raw = re.sub(r"<[^>]+>", "", raw)
        text = html.unescape(raw).strip()
        text = re.sub(r"\s+", " ", text)
        if text:
            paragraphs.append(text)
    return paragraphs


def chapter_number(raw: str) -> str:
    return raw.strip().replace(".", "")


def classify_function(text: str) -> str:
    t = text.lower()
    if any(k in t for k in ["geograf", "ubicación", "ciudad", "río", "montes", "pueblos", "tarteso", "córdoba", "mérida", "talavera"]):
        return "Autoridad geográfica/toponímica"
    if any(k in t for k in ["guerra", "batalla", "cartagin", "romanos", "escip", "numancia", "viriato", "sertorio"]):
        return "Autoridad narrativa bélica"
    if any(k in t for k in ["fabul", "fals", "desacredit", "dud", "contradic", "no sé", "cuidado"]):
        return "Fuente problemática o usada críticamente"
    if any(k in t for k in ["cristo", "apóstol", "cristian", "concilio", "obispo", "herej", "mártir", "santa", "igles"]):
        return "Autoridad eclesiástica/cristiana"
    if any(k in t for k in ["rey", "reyes", "fundación", "origen", "mitológ", "hércules", "teucro", "ulises"]):
        return "Autoridad sobre orígenes, genealogía o materia mítica"
    if any(k in t for k in ["era", "año", "cronolog", "fecha"]):
        return "Autoridad cronológica"
    if any(k in t for k in ["obra", "public", "rescat", "tradu", "manuscrit", "fragment"]):
        return "Autoridad bibliográfica o de transmisión textual"
    return "Autoridad historiográfica de apoyo"


def observation(text: str) -> str:
    notes = ["Borrador desde notas Word; pendiente relocalizar en 1601 y contrastar con 1617 solo si hace falta OCR."]
    t = text.lower()
    if "no sé" in t or "creo" in t or "quizás" in t or "posiblemente" in t:
        notes.append("Identificación insegura en las notas.")
    if "desacredit" in t or "fals" in t or "dud" in t:
        notes.append("Caso crítico/problemático.")
    if "margen" in t or "pág" in t:
        notes.append("Revisar página/margen indicado en las notas.")
    return " ".join(notes)


def normalize_text(text: str) -> str:
    return re.sub(r"\s+", " ", text.lower()).strip()


def detect_authors(text: str) -> list[tuple[int, str, str, str, str]]:
    hits = []
    for pattern, norm, work, kind in AUTHORS:
        for m in re.finditer(pattern, text, flags=re.IGNORECASE):
            hits.append((m.start(), m.group(0), norm, work, kind))
            break
    hits.sort(key=lambda x: x[0])
    seen = set()
    unique = []
    for hit in hits:
        key = hit[2]
        if key not in seen:
            unique.append(hit)
            seen.add(key)
    return unique


def detect_citation_authors(text: str) -> list[tuple[int, str, str, str, str]]:
    """Return likely Mariana-citation authors, not every name in an explanatory note.

    The Word notes often explain one cited author by comparing that author with
    other writers. A naive full-paragraph match overcounts those comparison names.
    This keeps the ordered table conservative: normally only names at the start of
    the note are treated as citation rows, while explicit list cues keep all names.
    """
    hits = detect_authors(text)
    if not hits:
        return []

    explicit_list = re.match(
        r"(?:Nueva\s+mención|Mención|Vuelve\s+a|De\s+nuevo|Junto\s+a|Todos\s+ellos)",
        text,
        flags=re.IGNORECASE,
    )
    if explicit_list:
        return hits

    first_pos = hits[0][0]
    if first_pos > 35:
        return []

    # Keep initial author lists, but avoid later explanatory comparisons.
    lead_limit = 170
    selected = [hit for hit in hits if hit[0] <= lead_limit]

    # One known note starts with Casiodoro and later joins Tertuliano as a paired citation.
    if re.match(r"Flavio\s+Magno\s+Aurelio\s+Casiodoro", text, flags=re.IGNORECASE):
        selected = [hit for hit in hits if hit[2] in {"Casiodoro", "Tertuliano", "Paulo Orosio"}]

    return selected


def build_rows() -> list[list[str]]:
    paragraphs = extract_paragraphs(DOCX)
    rows = []
    current_book = ""
    current_chapter = ""
    current_title = ""
    in_scope = False

    for paragraph in paragraphs:
        book_match = re.search(r"Libro\s+(Primero|Segundo|Tercero|Cuarto)", paragraph, flags=re.IGNORECASE)
        if book_match:
            current_book = BOOKS[book_match.group(1).capitalize()]
            current_chapter = ""
            current_title = ""
            in_scope = True
            continue

        if not in_scope:
            continue

        chapter_match = re.match(r"Cap\s+([IVXLCDM]+)\.\s*(.*)", paragraph, flags=re.IGNORECASE)
        if chapter_match:
            current_chapter = chapter_number(chapter_match.group(1))
            current_title = chapter_match.group(2).strip()
            continue

        if not current_book or not current_chapter:
            continue

        hits = detect_citation_authors(paragraph)
        for _pos, cited, norm_name, work, kind in hits:
            rows.append([
                len(rows) + 1,
                current_book,
                current_chapter,
                current_title,
                cited,
                norm_name,
                "",
                paragraph,
                classify_function(paragraph),
                kind,
                observation(paragraph),
            ])
    return rows


def write_csv(headers: list[str], rows: list[list[str]]) -> None:
    with OUT_CSV.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        writer.writerows(rows)


def write_xlsx(headers: list[str], rows: list[list[str]]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Usos autores I-IV"
    ws.append(headers)
    for row in rows:
        ws.append(row)

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = Border(top=thin, bottom=thin, left=thin, right=thin)

    widths = {
        "A": 8,
        "B": 8,
        "C": 10,
        "D": 38,
        "E": 28,
        "F": 30,
        "G": 36,
        "H": 72,
        "I": 42,
        "J": 32,
        "K": 58,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    for idx in range(2, ws.max_row + 1):
        ws.row_dimensions[idx].height = 90
    ws.row_dimensions[1].height = 35
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    meta = wb.create_sheet("Criterios")
    for row in [
        ["Campo", "Criterio"],
        ["Alcance", "Tomo I; Libros I-IV inclusive."],
        ["Base", "Documento Word original de notas del usuario."],
        ["Estado", "Borrador estructurado. Cada fila debe relocalizarse en 1601; la edición de 1617 solo sirve como apoyo OCR."],
        ["Obra asociada", "Se deja vacía por defecto. Las obras incluidas en las notas proceden de una búsqueda inicial y pueden no corresponder al texto de Mariana. Solo debe rellenarse tras revisión profunda del pasaje en 1601/1617."],
        ["Orden", "Respeta Libro > Capítulo > aparición en las notas."],
        ["Repeticiones", "Mantiene autores repetidos cuando aparecen en contextos o funciones distintas."],
    ]:
        meta.append(row)
    for row in meta.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = Border(top=thin, bottom=thin, left=thin, right=thin)
    for cell in meta[1]:
        cell.fill = header_fill
        cell.font = header_font
    meta.column_dimensions["A"].width = 18
    meta.column_dimensions["B"].width = 110

    wb.save(OUT_XLSX)


def main() -> None:
    headers = [
        "Orden",
        "Libro",
        "Capítulo",
        "Título del capítulo",
        "Autor citado",
        "Forma normalizada",
        "Obra asociada",
        "Pasaje / contexto",
        "Función de la cita",
        "Tipo de fuente",
        "Observaciones",
    ]
    rows = build_rows()
    write_csv(headers, rows)
    write_xlsx(headers, rows)
    print(f"created {OUT_XLSX} and {OUT_CSV} with {len(rows)} rows")


if __name__ == "__main__":
    main()
